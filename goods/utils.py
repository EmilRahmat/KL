import openpyxl
from goods.models import Variation, ProcessedReceipt, Products, Categories
from django.db import transaction
import logging

# Обновление по остаткам (старый способ)
def update_stock_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        sku = row[0]
        quantity = row[6]

        try:
            variation = Variation.objects.get(sku=sku)
            variation.quantity = quantity
            variation.save()
            print(f'✅ Остаток для {sku} обновлён на {quantity}')
        except Variation.DoesNotExist:
            print(f'⚠️ Вариация с SKU {sku} не найдена.')


# Обновление по чекам (новый способ)
def update_stock_by_receipts(file_path):
    

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    receipt_id = None
    items_to_process = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        col_2 = str(row[2]).strip() if row[2] else ""
        sku = str(row[10]).strip() if row[10] else ""
        qty = row[12]

        if col_2.startswith("Чек"):
            # Обработка предыдущего чека
            if receipt_id and items_to_process:
                if ProcessedReceipt.objects.filter(receipt_id=receipt_id).exists():
                    print(f'🔁 Чек {receipt_id} уже обработан — пропускаем.')
                else:
                    try:
                        with transaction.atomic():
                            for sku_val, qty_val in items_to_process:
                                try:
                                    variation = Variation.objects.get(sku=sku_val)
                                    variation.quantity = max(0, variation.quantity - int(qty_val))
                                    variation.save()
                                    print(f'🛒 {qty_val} списано из {sku_val}, осталось {variation.quantity}')
                                except Variation.DoesNotExist:
                                    print(f'❌ Вариация с SKU {sku_val} не найдена.')
                            ProcessedReceipt.objects.create(receipt_id=receipt_id)
                    except Exception as e:
                        print(f'⚠️ Ошибка при обработке чека {receipt_id}: {e}')

            # Начинаем новый чек
            receipt_id = col_2
            items_to_process = []

        elif receipt_id and sku and qty:
            items_to_process.append((sku, qty))

    # Обработка последнего чека (если есть)
    if receipt_id and items_to_process:
        if not ProcessedReceipt.objects.filter(receipt_id=receipt_id).exists():
            try:
                with transaction.atomic():
                    for sku_val, qty_val in items_to_process:
                        try:
                            variation = Variation.objects.get(sku=sku_val)
                            variation.quantity = max(0, variation.quantity - int(qty_val))
                            variation.save()
                            print(f'🛒 {qty_val} списано из {sku_val}, осталось {variation.quantity}')
                        except Variation.DoesNotExist:
                            print(f'❌ Вариация с SKU {sku_val} не найдена.')
                    ProcessedReceipt.objects.create(receipt_id=receipt_id)
            except Exception as e:
                print(f'⚠️ Ошибка при обработке последнего чека {receipt_id}: {e}')
        else:
            print(f'🔁 Последний чек {receipt_id} уже обработан — пропускаем.')

logger = logging.getLogger(__name__)

SEASON_MAPPING = {
    'all season': 'all_season',
    'осень-зима': 'autumn_winter',
    'весна-лето': 'spring_summer',
}

COLOR_MAPPING = {
    'белый': 'white',
    'черный': 'black',
    'красный': 'red',
    'бежевый': 'beige',
    'хаки': 'khaki',
    'розовый': 'ping',
    'фиолетовый': 'viol',
    'синий': 'blue',
    'зеленый': 'green',
    'желтый': 'yellow',
    'серый': 'grey',
    'коричневый': 'brown',
    'бордовый': 'burgundy',
}

def import_products_from_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    logs = []

    SEASON_MAPPING = {
        'all season': 'all_season',
        'spring summer': 'spring_summer',
        'autumn winter': 'autumn_winter',
        'весна-лето': 'spring_summer',
        'осень-зима': 'autumn_winter',
        'на любой сезон': 'all_season',
    }

    COLOR_MAPPING = {
        'белый': 'white',
        'черный': 'black',
        'синий': 'blue',
        'красный': 'red',
        'бежевый': 'beige',
        'хаки': 'khaki',
        'розовый': 'pink',
        'фиолетовый': 'violet',
        'зеленый': 'green',
        'желтый': 'yellow',
        'серый': 'grey',
        'коричневый': 'brown',
        'бордовый': 'burgundy',
    }

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) < 10:
            continue

        product_sku, name, price, category_id, var_sku, size, color, prod_size, qty, season, *_ = row

        if not product_sku or not var_sku:
            continue

        product_sku = str(product_sku).strip().replace(' ', '')
        var_sku = str(var_sku).strip().replace(' ', '')
        price = float(str(price).replace(' ', '').replace(',', '.'))
        qty = int(qty) if qty else 0

        season = str(season).strip().lower().replace('-', ' ')
        season = SEASON_MAPPING.get(season, 'all_season')

        color = str(color).strip().lower()
        color = COLOR_MAPPING.get(color, None)

        category = Categories.objects.filter(id=category_id).first()
        if not category:
            logs.append(f'⛔ Категория ID={category_id} не найдена. Пропущен SKU={product_sku}')
            continue

        product, created = Products.objects.get_or_create(
            sku=product_sku,
            defaults={
                'name': name,
                'price': price,
                'category': category,
                'season': season,
            }
        )

        if not created:
            updates = []
            if product.name != name:
                updates.append(f"название '{product.name}' → '{name}'")
                product.name = name
            if product.price != price:
                updates.append(f"цена {product.price} → {price}")
                product.price = price
            if product.category_id != category.id:
                updates.append(f"категория ID {product.category_id} → {category.id}")
                product.category = category
            if product.season != season:
                updates.append(f"сезон {product.season} → {season}")
                product.season = season
            if updates:
                product.save()
                logs.append(f"🔄 Обновлён товар [{product.sku}]: " + "; ".join(updates))

        variation, v_created = Variation.objects.update_or_create(
            sku=var_sku,
            defaults={
                'product': product,
                'size': size,
                'color': color,
                'producer_size': prod_size,
                'quantity': qty,
                'is_active': True
            }
        )

        if v_created:
            logs.append(f"➕ Создана вариация {var_sku}")
        else:
            logs.append(f"🔁 Обновлена вариация {var_sku}")

    return logs
